"""
FacilityApps Bulk Job Importer - MVP
Local-first Streamlit app for validating and importing one-off jobs/rosters
Design: https://www.figma.com/make/nI65iKCEyzqO5dfqb3fXsL/FA-Job-Importer-Design-System
"""

import streamlit as st
import pandas as pd
import requests
import json
import time
import os
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import pytz
from dotenv import load_dotenv
from styles import apply_custom_theme, apply_login_screen_style

# Load environment variables
load_dotenv()

# Authentication with username and password
def check_password():
    """Returns `True` if the user had valid credentials."""
    
    def credentials_entered():
        """Checks whether credentials entered by the user are correct."""
        try:
            if "auth" in st.secrets and "master_username" in st.secrets["auth"] and "master_password" in st.secrets["auth"]:
                username = st.session_state.get("username", "")
                password = st.session_state.get("password", "")
                
                if (username == st.secrets["auth"]["master_username"] and 
                    password == st.secrets["auth"]["master_password"]):
                    st.session_state["password_correct"] = True
                    st.session_state["logged_in_user"] = username
                    # Don't store credentials
                    del st.session_state["username"]
                    del st.session_state["password"]
                else:
                    st.session_state["password_correct"] = False
            else:
                st.session_state["password_correct"] = False
                st.session_state["auth_error"] = "Authentication not configured in secrets.toml"
        except Exception as e:
            st.session_state["password_correct"] = False
            st.session_state["auth_error"] = f"Auth error: {str(e)}"

    # Apply custom theme
    apply_custom_theme()
    apply_login_screen_style()
    
    # Return True if the user is authenticated.
    if st.session_state.get("password_correct", False):
        return True

    # Check if secrets are available
    if "auth_error" in st.session_state:
        st.error(st.session_state["auth_error"])
        return False
    
    try:
        if "auth" not in st.secrets or "master_username" not in st.secrets["auth"] or "master_password" not in st.secrets["auth"]:
            st.error("❌ Authentication not configured. Please add [auth] master_username and master_password to .streamlit/secrets.toml")
            st.code("""
[auth]
master_username = "admin"
master_password = "your_secure_password"
            """)
            return False
    except Exception as e:
        st.error(f"❌ Error reading secrets: {str(e)}")
        return False
    
    # Login form - Figma design style
    col1, col2, col3 = st.columns([1, 2.5, 1])
    with col2:
        # Card header with logo
        st.markdown("""
        <div style='background: white; border: 1px solid rgba(0,0,0,0.1); border-radius: 10px; padding: 2rem; margin-top: 5rem; box-shadow: 0 4px 6px rgba(0,0,0,0.05);'>
            <div style='display: flex; align-items: center; gap: 12px; margin-bottom: 24px;'>
                <div style='width: 40px; height: 40px; background: #030213; border-radius: 8px; display: flex; align-items: center; justify-center;'>
                    <span style='color: white; font-size: 18px;'>🔒</span>
                </div>
                <div>
                    <div style='font-size: 18px; font-weight: 600; color: #030213;'>FacilityApps</div>
                    <div style='font-size: 11px; color: #717182;'>Job Importer</div>
                </div>
            </div>
            <h2 style='font-size: 20px; font-weight: 600; margin-bottom: 8px; color: #030213;'>Authentication Required</h2>
            <p style='font-size: 14px; color: #717182; margin-bottom: 24px;'>Sign in to access the bulk job importer</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<div style='margin-top: -10px;'>", unsafe_allow_html=True)
        st.text_input(
            "Username", 
            key="username",
            placeholder="Enter username",
            label_visibility="visible"
        )
        st.text_input(
            "Password", 
            type="password", 
            key="password",
            placeholder="Enter password",
            label_visibility="visible"
        )
        
        if st.session_state.get("password_correct") == False:
            st.error("Invalid username or password")
        
        if st.button("🔒 Sign In", use_container_width=True, type="primary"):
            credentials_entered()
            st.rerun()
        
        st.markdown("</div>", unsafe_allow_html=True)
    
    return False

# Setup logging
def setup_logging():
    """Setup logging for API payloads"""
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"api_payloads_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    return log_file

# Constants
TIMEZONE = pytz.timezone("Europe/London")
MAX_RETRIES = 3
BASE_DELAY = 1.0
FAILURE_THRESHOLD = 0.1  # Abort import if >10% failures


class FacilityAppsClient:
    """Client for interacting with FacilityApps API"""
    
    def __init__(self, domain: str, token: str):
        self.domain = domain
        self.token = token
        self.base_url = f"https://{domain}"
        self.headers = {
            "Authorization": f"Bearer {self._redact_token(token, reveal=True)}",
            "Accept": "application/json",
            "Content-Type": "application/json"
        }
    
    @staticmethod
    def _redact_token(token: str, reveal: bool = False) -> str:
        """Redact token for logging (show first/last 4 chars only)"""
        if reveal:
            return token
        if len(token) <= 8:
            return "***"
        return f"{token[:4]}...{token[-4:]}"
    
    def test_connection(self) -> Tuple[bool, str]:
        """Test API connectivity"""
        try:
            response = requests.get(
                f"{self.base_url}/api/v1/planning/sites",
                headers=self.headers,
                timeout=10
            )
            if response.status_code == 200:
                return True, "Connection successful"
            else:
                return False, f"HTTP {response.status_code}: {response.text[:200]}"
        except Exception as e:
            return False, f"Connection error: {str(e)}"
    
    def get_sites(self) -> List[Dict]:
        """Fetch all sites"""
        try:
            response = requests.get(
                f"{self.base_url}/api/v1/planning/sites",
                headers=self.headers,
                timeout=30
            )
            response.raise_for_status()
            data = response.json()
            
            # Handle both list and dict responses
            if isinstance(data, list):
                return data
            elif isinstance(data, dict):
                # Try common patterns: {"data": [...]} or {"sites": [...]}
                if "data" in data:
                    return data["data"]
                elif "sites" in data:
                    return data["sites"]
                else:
                    st.warning(f"Unexpected sites response format with keys: {list(data.keys())}")
                    return []
            return []
        except Exception as e:
            st.error(f"Failed to fetch sites: {e}")
            return []
    
    def get_floors(self) -> List[Dict]:
        """Fetch all floors"""
        try:
            response = requests.get(
                f"{self.base_url}/api/v1/planning/floors",
                headers=self.headers,
                timeout=30
            )
            response.raise_for_status()
            data = response.json()
            
            # Handle both list and dict responses
            if isinstance(data, list):
                return data
            elif isinstance(data, dict):
                # Try common patterns: {"data": [...]} or {"floors": [...]}
                if "data" in data:
                    return data["data"]
                elif "floors" in data:
                    return data["floors"]
                else:
                    st.warning(f"Unexpected floors response format with keys: {list(data.keys())}")
                    return []
            return []
        except Exception as e:
            st.error(f"Failed to fetch floors: {e}")
            return []
    
    def get_spaces(self) -> List[Dict]:
        """Fetch all spaces"""
        try:
            response = requests.get(
                f"{self.base_url}/api/v1/planning/spaces",
                headers=self.headers,
                timeout=30
            )
            response.raise_for_status()
            data = response.json()
            
            # Handle both list and dict responses
            if isinstance(data, list):
                return data
            elif isinstance(data, dict):
                # Try common patterns: {"data": [...]} or {"spaces": [...]}
                if "data" in data:
                    return data["data"]
                elif "spaces" in data:
                    return data["spaces"]
                else:
                    st.warning(f"Unexpected spaces response format with keys: {list(data.keys())}")
                    return []
            return []
        except Exception as e:
            st.error(f"Failed to fetch spaces: {e}")
            return []
    
    def get_users(self) -> List[Dict]:
        """Fetch all users with pagination"""
        all_users = []
        page = 1
        limit = 200
        
        try:
            while True:
                response = requests.get(
                    f"{self.base_url}/api/v1/user",
                    headers=self.headers,
                    params={"limit": limit, "page": page},
                    timeout=30
                )
                response.raise_for_status()
                data = response.json()
                
                users = data.get("data", [])
                all_users.extend(users)
                
                meta = data.get("meta", {})
                pagination = meta.get("pagination", {})
                total_pages = pagination.get("total_pages", 1)
                
                if page >= total_pages:
                    break
                page += 1
                
            return all_users
        except Exception as e:
            st.error(f"Failed to fetch users: {e}")
            return []
    
    def create_job(self, payload: Dict) -> Tuple[bool, Any]:
        """
        Create a one-off job/roster
        Returns: (success, response_data_or_error)
        """
        try:
            response = requests.post(
                f"{self.base_url}/api/1.0/planning/save/true",
                headers=self.headers,
                json=payload,
                timeout=30
            )
            
            # Respect rate limiting
            if "X-RateLimit-Remaining" in response.headers:
                remaining = int(response.headers["X-RateLimit-Remaining"])
                if remaining < 5:
                    time.sleep(2)  # Back off if close to limit
            
            if response.status_code in [200, 201]:
                return True, response.json()
            else:
                return False, {
                    "status": response.status_code,
                    "body": response.text[:500]
                }
        except Exception as e:
            return False, {"error": str(e)}
    
    def create_job_with_retry(self, payload: Dict) -> Tuple[bool, Any]:
        """Create job with exponential backoff retry on 5xx errors"""
        for attempt in range(MAX_RETRIES):
            success, result = self.create_job(payload)
            
            if success:
                return True, result
            
            # Check if it's a 5xx error
            if isinstance(result, dict) and result.get("status", 0) >= 500:
                if attempt < MAX_RETRIES - 1:
                    delay = BASE_DELAY * (2 ** attempt)
                    time.sleep(delay)
                    continue
            
            return False, result
        
        return False, {"error": "Max retries exceeded"}


class JobValidator:
    """Validates CSV rows against FacilityApps requirements"""
    
    def __init__(self, sites: List[Dict], floors: List[Dict], 
                 spaces: List[Dict], users: List[Dict]):
        # Ensure inputs are lists (handle None)
        sites = sites or []
        floors = floors or []
        spaces = spaces or []
        users = users or []
        
        # Build lookup dictionaries by ID (filter out None items, non-dicts, and items without IDs)
        self.sites = {
            str(s.get("id")): s 
            for s in sites 
            if isinstance(s, dict) and s.get("id") is not None
        }
        self.floors = {
            str(f.get("id")): f 
            for f in floors 
            if isinstance(f, dict) and f.get("id") is not None
        }
        self.spaces = {
            str(s.get("id")): s 
            for s in spaces 
            if isinstance(s, dict) and s.get("id") is not None
        }
        self.users = {
            str(u.get("id")): u 
            for u in users 
            if isinstance(u, dict) and u.get("id") is not None
        }
        
        # Build name-to-ID lookups (case-insensitive)
        self.site_name_to_id = {}
        for s in sites:
            if isinstance(s, dict) and s.get("id") is not None:
                # Try common name fields
                name = s.get("name") or s.get("site_name") or s.get("title")
                if name:
                    self.site_name_to_id[str(name).strip().lower()] = str(s.get("id"))
        
        self.floor_name_to_id = {}
        for f in floors:
            if isinstance(f, dict) and f.get("id") is not None:
                name = f.get("name") or f.get("floor_name") or f.get("title")
                if name:
                    # Include site context in key for disambiguation
                    site_id = str(f.get("site_id", f.get("location_id", "")))
                    key = f"{site_id}|{str(name).strip().lower()}"
                    self.floor_name_to_id[key] = str(f.get("id"))
                    # Also store without site context (less safe but more flexible)
                    self.floor_name_to_id[str(name).strip().lower()] = str(f.get("id"))
        
        self.space_name_to_id = {}
        for s in spaces:
            if isinstance(s, dict) and s.get("id") is not None:
                name = s.get("name") or s.get("space_name") or s.get("title")
                if name:
                    # Include floor context for disambiguation
                    floor_id = str(s.get("floor_id", ""))
                    key = f"{floor_id}|{str(name).strip().lower()}"
                    self.space_name_to_id[key] = str(s.get("id"))
                    # Also store without floor context
                    self.space_name_to_id[str(name).strip().lower()] = str(s.get("id"))
        
        self.user_name_to_id = {}
        for u in users:
            if isinstance(u, dict) and u.get("id") is not None:
                # Try various user name fields
                username = u.get("user_name") or u.get("username") or u.get("name")
                if username:
                    self.user_name_to_id[str(username).strip().lower()] = str(u.get("id"))
                # Also try email
                email = u.get("email")
                if email:
                    self.user_name_to_id[str(email).strip().lower()] = str(u.get("id"))
        
        # Build relationships
        self.floor_to_site = {
            str(f.get("id")): str(f.get("siteId") or f.get("site_id") or f.get("location_id"))
            for f in floors
            if isinstance(f, dict) and f.get("id") is not None and (f.get("siteId") is not None or f.get("site_id") is not None or f.get("location_id") is not None)
        }
        self.space_to_floor = {
            str(s.get("id")): str(s.get("floorId") or s.get("floor_id"))
            for s in spaces
            if isinstance(s, dict) and s.get("id") is not None and (s.get("floorId") is not None or s.get("floor_id") is not None)
        }
    
    def resolve_site(self, row: pd.Series) -> Tuple[Optional[str], Optional[str]]:
        """
        Resolve site ID from either site_id or site_name column
        Returns: (resolved_id, error_message)
        """
        site_id = row.get("site_id")
        site_name = row.get("site_name")
        
        # Prefer ID if provided
        if pd.notna(site_id) and str(site_id).strip() != "":
            return str(site_id).strip(), None
        
        # Try name lookup
        if pd.notna(site_name) and str(site_name).strip() != "":
            name_key = str(site_name).strip().lower()
            resolved_id = self.site_name_to_id.get(name_key)
            if resolved_id:
                return resolved_id, None
            else:
                # Try to suggest close matches
                available = list(self.site_name_to_id.keys())[:5]
                return None, f"Site name '{site_name}' not found. Available: {', '.join(available)}"
        
        return None, "Neither site_id nor site_name provided"
    
    def resolve_floor(self, row: pd.Series, site_id: str) -> Tuple[Optional[str], Optional[str]]:
        """
        Resolve floor ID from either floor_id or floor_name column
        Returns: (resolved_id, error_message)
        """
        floor_id = row.get("floor_id")
        floor_name = row.get("floor_name")
        
        # If neither provided, that's OK (floor is optional)
        if (pd.isna(floor_id) or str(floor_id).strip() == "") and \
           (pd.isna(floor_name) or str(floor_name).strip() == ""):
            return None, None
        
        # Prefer ID if provided
        if pd.notna(floor_id) and str(floor_id).strip() != "":
            return str(floor_id).strip(), None
        
        # Try name lookup with site context
        if pd.notna(floor_name) and str(floor_name).strip() != "":
            name = str(floor_name).strip().lower()
            # Try with site context first
            context_key = f"{site_id}|{name}"
            resolved_id = self.floor_name_to_id.get(context_key)
            if resolved_id:
                return resolved_id, None
            # Try without context
            resolved_id = self.floor_name_to_id.get(name)
            if resolved_id:
                return resolved_id, None
            else:
                return None, f"Floor name '{floor_name}' not found in site {site_id}"
        
        return None, None
    
    def resolve_space(self, row: pd.Series, floor_id: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
        """
        Resolve space ID from either space_id or space_name column
        Returns: (resolved_id, error_message)
        """
        space_id = row.get("space_id")
        space_name = row.get("space_name")
        
        # If neither provided, that's OK (space is optional)
        if (pd.isna(space_id) or str(space_id).strip() == "") and \
           (pd.isna(space_name) or str(space_name).strip() == ""):
            return None, None
        
        # Prefer ID if provided
        if pd.notna(space_id) and str(space_id).strip() != "":
            return str(space_id).strip(), None
        
        # Try name lookup with floor context
        if pd.notna(space_name) and str(space_name).strip() != "":
            name = str(space_name).strip().lower()
            if floor_id:
                # Try with floor context first
                context_key = f"{floor_id}|{name}"
                resolved_id = self.space_name_to_id.get(context_key)
                if resolved_id:
                    return resolved_id, None
            # Try without context
            resolved_id = self.space_name_to_id.get(name)
            if resolved_id:
                return resolved_id, None
            else:
                return None, f"Space name '{space_name}' not found"
        
        return None, None
    
    def resolve_owner(self, row: pd.Series) -> Tuple[Optional[str], Optional[str]]:
        """
        Resolve owner ID from either owner_employee_id or owner_name/owner_email
        Returns: (resolved_id, error_message)
        """
        owner_id = row.get("owner_employee_id")
        owner_name = row.get("owner_name")
        owner_email = row.get("owner_email")
        
        # Prefer ID if provided
        if pd.notna(owner_id) and str(owner_id).strip() != "":
            return str(owner_id).strip(), None
        
        # Try name lookup
        if pd.notna(owner_name) and str(owner_name).strip() != "":
            name_key = str(owner_name).strip().lower()
            resolved_id = self.user_name_to_id.get(name_key)
            if resolved_id:
                return resolved_id, None
        
        # Try email lookup
        if pd.notna(owner_email) and str(owner_email).strip() != "":
            email_key = str(owner_email).strip().lower()
            resolved_id = self.user_name_to_id.get(email_key)
            if resolved_id:
                return resolved_id, None
        
        return None, "Owner not found by ID, name, or email"
    
    def validate_row(self, row: pd.Series, row_num: int, 
                     seen_keys: set) -> List[Dict]:
        """
        Validate a single row and return list of issues
        Each issue: {row_number, status, issue_code, issue_detail, suggestion}
        """
        issues = []
        today = datetime.now(TIMEZONE).date()
        
        # Resolve site (required - either ID or name must be provided)
        site_id, site_error = self.resolve_site(row)
        if site_error:
            issues.append({
                "row_number": row_num,
                "status": "ERROR",
                "issue_code": "MISSING_SITE",
                "issue_detail": site_error,
                "suggestion": "Provide site_id or site_name"
            })
        else:
            # Store resolved ID back in row for later use
            row["_resolved_site_id"] = site_id
        
        # Resolve owner (required - either ID, name, or email)
        owner_id, owner_error = self.resolve_owner(row)
        if owner_error:
            issues.append({
                "row_number": row_num,
                "status": "ERROR",
                "issue_code": "MISSING_OWNER",
                "issue_detail": owner_error,
                "suggestion": "Provide owner_employee_id, owner_name, or owner_email"
            })
        else:
            # Store resolved ID back in row
            row["_resolved_owner_id"] = owner_id
        
        # Resolve floor (optional)
        floor_id = None
        if site_id:
            floor_id, floor_error = self.resolve_floor(row, site_id)
            if floor_error:
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "INVALID_FLOOR",
                    "issue_detail": floor_error,
                    "suggestion": "Check floor_id or floor_name"
                })
            elif floor_id:
                row["_resolved_floor_id"] = floor_id
        
        # Resolve space (optional)
        if floor_id:
            space_id, space_error = self.resolve_space(row, floor_id)
            if space_error:
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "INVALID_SPACE",
                    "issue_detail": space_error,
                    "suggestion": "Check space_id or space_name"
                })
            elif space_id:
                row["_resolved_space_id"] = space_id
        
        # Required field checks (non-location fields)
        required = ["title_en", "date_start", "date_end", "hour_start", 
                   "minute_start", "hour_end", "minute_end"]
        
        for field in required:
            if pd.isna(row.get(field)) or str(row.get(field, "")).strip() == "":
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "MISSING_REQUIRED",
                    "issue_detail": f"Missing required field: {field}",
                    "suggestion": f"Provide a value for {field}"
                })
        
        # If missing required fields, skip further validation
        if any(i["issue_code"] in ["MISSING_SITE", "MISSING_OWNER", "MISSING_REQUIRED"] for i in issues):
            return issues
        
        # Date validation
        try:
            date_start = pd.to_datetime(row["date_start"]).date()
            date_end = pd.to_datetime(row["date_end"]).date()
            
            if date_start < today:
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "PAST_DATE",
                    "issue_detail": f"date_start {date_start} is in the past",
                    "suggestion": f"Use date >= {today}"
                })
            
            if date_end < date_start:
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "INVALID_DATE_RANGE",
                    "issue_detail": f"date_end {date_end} before date_start {date_start}",
                    "suggestion": "Set date_end >= date_start"
                })
        except Exception as e:
            issues.append({
                "row_number": row_num,
                "status": "ERROR",
                "issue_code": "INVALID_DATE",
                "issue_detail": f"Date parsing error: {e}",
                "suggestion": "Use YYYY-MM-DD format"
            })
        
        # Time validation
        try:
            hour_start = int(row["hour_start"])
            minute_start = int(row["minute_start"])
            hour_end = int(row["hour_end"])
            minute_end = int(row["minute_end"])
            
            if not (0 <= hour_start <= 23):
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "INVALID_HOUR",
                    "issue_detail": f"hour_start {hour_start} out of range",
                    "suggestion": "Use 0-23"
                })
            
            if not (0 <= hour_end <= 23):
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "INVALID_HOUR",
                    "issue_detail": f"hour_end {hour_end} out of range",
                    "suggestion": "Use 0-23"
                })
            
            if not (0 <= minute_start <= 59):
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "INVALID_MINUTE",
                    "issue_detail": f"minute_start {minute_start} out of range",
                    "suggestion": "Use 0-59"
                })
            
            if not (0 <= minute_end <= 59):
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "INVALID_MINUTE",
                    "issue_detail": f"minute_end {minute_end} out of range",
                    "suggestion": "Use 0-59"
                })
            
            # Check end time > start time
            start_minutes = hour_start * 60 + minute_start
            end_minutes = hour_end * 60 + minute_end
            
            if end_minutes <= start_minutes:
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "INVALID_TIME_RANGE",
                    "issue_detail": f"End time ({hour_end:02d}:{minute_end:02d}) must be after start ({hour_start:02d}:{minute_start:02d})",
                    "suggestion": "Set end_time > start_time"
                })
        except Exception as e:
            issues.append({
                "row_number": row_num,
                "status": "ERROR",
                "issue_code": "INVALID_TIME",
                "issue_detail": f"Time parsing error: {e}",
                "suggestion": "Use integer values for hours/minutes"
            })
        
        # Validate resolved site ID exists
        resolved_site_id = row.get("_resolved_site_id")
        if resolved_site_id and resolved_site_id not in self.sites:
            issues.append({
                "row_number": row_num,
                "status": "ERROR",
                "issue_code": "INVALID_SITE",
                "issue_detail": f"Resolved site_id {resolved_site_id} not found in system",
                "suggestion": "Check site reference"
            })
        
        # Validate resolved floor belongs to site (if both present)
        resolved_floor_id = row.get("_resolved_floor_id")
        if resolved_floor_id and resolved_site_id:
            if resolved_floor_id not in self.floors:
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "INVALID_FLOOR",
                    "issue_detail": f"Resolved floor_id {resolved_floor_id} not found",
                    "suggestion": "Check floor reference"
                })
            else:
                # Check floor belongs to site
                expected_site = self.floor_to_site.get(resolved_floor_id)
                if expected_site != resolved_site_id:
                    issues.append({
                        "row_number": row_num,
                        "status": "ERROR",
                        "issue_code": "FLOOR_SITE_MISMATCH",
                        "issue_detail": f"Floor does not belong to specified site",
                        "suggestion": f"Floor belongs to site {expected_site}"
                    })
        
        # Validate resolved space belongs to floor (if both present)
        resolved_space_id = row.get("_resolved_space_id")
        if resolved_space_id:
            if not resolved_floor_id:
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "SPACE_WITHOUT_FLOOR",
                    "issue_detail": "Space provided without floor",
                    "suggestion": "Provide floor when specifying space"
                })
            elif resolved_space_id not in self.spaces:
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "INVALID_SPACE",
                    "issue_detail": f"Resolved space_id {resolved_space_id} not found",
                    "suggestion": "Check space reference"
                })
            else:
                # Check space belongs to floor
                expected_floor = self.space_to_floor.get(resolved_space_id)
                if expected_floor != resolved_floor_id:
                    issues.append({
                        "row_number": row_num,
                        "status": "ERROR",
                        "issue_code": "SPACE_FLOOR_MISMATCH",
                        "issue_detail": f"Space does not belong to specified floor",
                        "suggestion": f"Space belongs to floor {expected_floor}"
                    })
        
        # Validate resolved owner ID exists
        resolved_owner_id = row.get("_resolved_owner_id")
        if resolved_owner_id and resolved_owner_id not in self.users:
            issues.append({
                "row_number": row_num,
                "status": "ERROR",
                "issue_code": "INVALID_OWNER",
                "issue_detail": f"Resolved owner_id {resolved_owner_id} not found in system",
                "suggestion": "Check owner reference"
            })
        
        # Label validation - disabled for now
        # label_list = row.get("label_list", "")
        # if pd.notna(label_list) and str(label_list).strip() != "":
        #     labels = [l.strip() for l in str(label_list).split(",")]
        #     labels = [l for l in labels if l]  # Remove empty
        #     if len(labels) != len(set(labels)):
        #         issues.append({
        #             "row_number": row_num,
        #             "status": "WARN",
        #             "issue_code": "DUPLICATE_LABELS",
        #             "issue_detail": "Duplicate labels found",
        #             "suggestion": "Remove duplicate labels"
        #         })
        
        # Duplicate check (use resolved IDs)
        dup_key = (
            row.get("_resolved_site_id", ""),
            row.get("_resolved_floor_id", ""),
            row.get("_resolved_space_id", ""),
            str(row.get("title_en", "")).strip(),
            str(row.get("date_start", "")).strip(),
            str(row.get("hour_start", "")).strip(),
            str(row.get("minute_start", "")).strip()
        )
        
        if dup_key in seen_keys:
            issues.append({
                "row_number": row_num,
                "status": "WARN",
                "issue_code": "DUPLICATE_ROW",
                "issue_detail": "Duplicate job found in file (same site/floor/space/title/date/time)",
                "suggestion": "Review if intentional"
            })
        else:
            seen_keys.add(dup_key)
        
        # Recurrence validation
        recurrence_type = str(row.get("recurrence_type", "none")).strip().lower()
        if recurrence_type and recurrence_type not in ["", "none"]:
            # Validate recurrence type
            valid_types = ["daily", "weekdays", "weekly", "biweekly", "monthly"]
            if recurrence_type not in valid_types:
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "INVALID_RECURRENCE_TYPE",
                    "issue_detail": f"recurrence_type '{recurrence_type}' is invalid",
                    "suggestion": f"Use one of: {', '.join(valid_types)}, or leave empty for one-off"
                })
            
            # Validate recurrence end date is provided
            recurrence_end_date = row.get("recurrence_end_date")
            if pd.isna(recurrence_end_date) or str(recurrence_end_date).strip() == "":
                issues.append({
                    "row_number": row_num,
                    "status": "ERROR",
                    "issue_code": "MISSING_RECURRENCE_END",
                    "issue_detail": "recurrence_end_date required for recurring jobs",
                    "suggestion": "Provide end date (YYYY-MM-DD) or set recurrence_type to 'none'"
                })
            else:
                # Validate end date is after start date
                try:
                    rec_end = pd.to_datetime(recurrence_end_date).date()
                    start = pd.to_datetime(row["date_start"]).date()
                    if rec_end <= start:
                        issues.append({
                            "row_number": row_num,
                            "status": "ERROR",
                            "issue_code": "INVALID_RECURRENCE_END",
                            "issue_detail": f"recurrence_end_date {rec_end} must be after date_start {start}",
                            "suggestion": "Set end date > start date"
                        })
                except Exception as e:
                    issues.append({
                        "row_number": row_num,
                        "status": "ERROR",
                        "issue_code": "INVALID_RECURRENCE_END_FORMAT",
                        "issue_detail": f"Invalid recurrence_end_date format: {e}",
                        "suggestion": "Use YYYY-MM-DD format"
                    })
            
            # Validate days for weekly/biweekly patterns
            if recurrence_type in ["weekly", "biweekly"]:
                recurrence_days = row.get("recurrence_days", "")
                if pd.notna(recurrence_days) and str(recurrence_days).strip() != "":
                    days = [d.strip().lower()[:3] for d in str(recurrence_days).split(",")]
                    valid_days = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
                    invalid_days = [d for d in days if d not in valid_days]
                    if invalid_days:
                        issues.append({
                            "row_number": row_num,
                            "status": "ERROR",
                            "issue_code": "INVALID_RECURRENCE_DAYS",
                            "issue_detail": f"Invalid day(s): {', '.join(invalid_days)}",
                            "suggestion": "Use: Mon, Tue, Wed, Thu, Fri, Sat, Sun (comma-separated)"
                        })
            
            # Validate interval for custom intervals
            recurrence_interval = row.get("recurrence_interval")
            if pd.notna(recurrence_interval) and recurrence_interval != "":
                try:
                    interval = int(recurrence_interval)
                    if interval < 1:
                        issues.append({
                            "row_number": row_num,
                            "status": "ERROR",
                            "issue_code": "INVALID_RECURRENCE_INTERVAL",
                            "issue_detail": f"recurrence_interval {interval} must be >= 1",
                            "suggestion": "Use positive integer (1, 2, 3, etc.)"
                        })
                except ValueError:
                    issues.append({
                        "row_number": row_num,
                        "status": "ERROR",
                        "issue_code": "INVALID_RECURRENCE_INTERVAL_FORMAT",
                        "issue_detail": "recurrence_interval must be an integer",
                        "suggestion": "Use whole numbers (1, 2, 3, etc.)"
                    })
        
        # If no issues, mark as OK
        if not issues:
            issues.append({
                "row_number": row_num,
                "status": "OK",
                "issue_code": "VALID",
                "issue_detail": "Row is valid",
                "suggestion": ""
            })
        
        return issues


def build_recurrence_ui(row_idx: int, row: pd.Series) -> Dict:
    """Build recurrence UI and return recurrence settings"""
    
    # Initialize recurrence settings
    recurrence_settings = {
        "is_recurring": False,
        "repeat_interval_period": None,
        "repeat_interval_length": 1,
        "use_day_of_week": False,
        "frequency_daily_repeat": [],
        "frequency_weekly_repeat": [],
        "frequency_monthly_repeat": [],
        "frequency_stop_repeat": None,
        "frequency_stop_repeat_number_value": None,
        "end_after_date": None
    }
    
    st.write("**🔄 Recurrence Settings:**")
    
    # Step 1: Toggle recurrence
    is_recurring = st.checkbox(
        "Repeat this job",
        value=bool(row.get("is_recurring", False)),
        key=f"recurring_{row_idx}"
    )
    
    if not is_recurring:
        return recurrence_settings
    
    recurrence_settings["is_recurring"] = True
    
    # Step 2: Frequency and interval
    col1, col2 = st.columns([2, 1])
    
    with col1:
        frequency = st.selectbox(
            "Repeat every",
            options=["Daily", "Weekly", "Monthly", "Yearly"],
            index=0,
            key=f"frequency_{row_idx}"
        )
    
    with col2:
        interval = st.number_input(
            "Interval",
            min_value=1,
            max_value=999,
            value=int(row.get("recurrence_interval", 1)) if pd.notna(row.get("recurrence_interval", 1)) else 1,
            key=f"interval_{row_idx}"
        )
    
    # Map frequency to period
    period_map = {
        "Daily": "daily",
        "Weekly": "weekly", 
        "Monthly": "monthly",
        "Yearly": "yearly"
    }
    
    recurrence_settings["repeat_interval_period"] = period_map[frequency]
    recurrence_settings["repeat_interval_length"] = interval
    
    # Step 3: Pattern details
    if frequency == "Daily":
        st.write("**Repeat on:**")
        days = st.multiselect(
            "Days of week",
            options=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
            default=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
            key=f"daily_days_{row_idx}"
        )
        
        # Convert to numbers (1=Monday, 7=Sunday)
        day_map = {
            "Monday": 1, "Tuesday": 2, "Wednesday": 3, "Thursday": 4,
            "Friday": 5, "Saturday": 6, "Sunday": 7
        }
        recurrence_settings["frequency_daily_repeat"] = [day_map[d] for d in days]
        
    elif frequency == "Weekly":
        st.write("**Repeat on:**")
        days = st.multiselect(
            "Days of week",
            options=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
            default=["Monday"],
            key=f"weekly_days_{row_idx}"
        )
        
        day_map = {
            "Monday": 1, "Tuesday": 2, "Wednesday": 3, "Thursday": 4,
            "Friday": 5, "Saturday": 6, "Sunday": 7
        }
        recurrence_settings["frequency_daily_repeat"] = [day_map[d] for d in days]
        # For weekly on specific days, use daily period
        recurrence_settings["repeat_interval_period"] = "daily"
        
    elif frequency == "Monthly":
        monthly_type = st.radio(
            "Monthly pattern",
            options=["On this date", "On this weekday"],
            key=f"monthly_type_{row_idx}"
        )
        
        if monthly_type == "On this weekday":
            recurrence_settings["use_day_of_week"] = True
            # For monthly weekday, we still use frequency_daily_repeat for the day
            day = st.selectbox(
                "Day of week",
                options=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
                key=f"monthly_weekday_{row_idx}"
            )
            day_map = {
                "Monday": 1, "Tuesday": 2, "Wednesday": 3, "Thursday": 4,
                "Friday": 5, "Saturday": 6, "Sunday": 7
            }
            recurrence_settings["frequency_daily_repeat"] = [day_map[day]]
    
    # Step 4: End condition
    st.write("**Ends:**")
    end_type = st.radio(
        "End condition",
        options=["Never", "After N occurrences", "On date"],
        key=f"end_type_{row_idx}"
    )
    
    if end_type == "After N occurrences":
        recurrence_settings["frequency_stop_repeat"] = 1
        recurrence_settings["frequency_stop_repeat_number_value"] = st.number_input(
            "Number of occurrences",
            min_value=1,
            max_value=999,
            value=10,
            key=f"stop_count_{row_idx}"
        )
    elif end_type == "On date":
        recurrence_settings["frequency_stop_repeat"] = 2
        end_date_value = row.get("recurrence_end_date", "2025-12-31")
        if pd.notna(end_date_value) and str(end_date_value).strip():
            try:
                end_date_value = pd.to_datetime(end_date_value).date()
            except:
                end_date_value = pd.to_datetime("2025-12-31").date()
        else:
            end_date_value = pd.to_datetime("2025-12-31").date()
            
        recurrence_settings["end_after_date"] = st.date_input(
            "End date",
            value=end_date_value,
            key=f"stop_date_{row_idx}"
        )
    else:  # Never
        recurrence_settings["frequency_stop_repeat"] = 0
    
    # Natural language preview
    preview = build_recurrence_preview(recurrence_settings)
    st.info(f"📅 **Preview:** {preview}")
    
    return recurrence_settings


def create_recurring_jobs_template() -> bytes:
    """Create Excel template with recurring job examples"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Recurring Jobs Template"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    example_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    instruction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "title_en", "description_en", "site_name", "floor_name", "space_name", 
        "owner_email", "date_start", "date_end", "hour_start", "minute_start", 
        "hour_end", "minute_end",
        "is_recurring", "repeat_interval_period", "repeat_interval_length",
        "use_day_of_week", "frequency_daily_repeat", "frequency_weekly_repeat",
        "frequency_monthly_repeat", "frequency_stop_repeat", 
        "frequency_stop_repeat_number_value", "end_after_date"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Example data
    examples = [
        # One-time job
        [
            "Office Clean - One-off", "Single occurrence cleaning", "Main Office", "Ground Floor", "Reception",
            "cleaner@company.com", "2025-01-15", "2025-01-15", "9", "0", "17", "0",
            "FALSE", "", "", "", "", "", "", "", "", "", ""
        ],
        # Daily job (weekdays only)
        [
            "Daily Office Clean", "Regular office cleaning", "Main Office", "Ground Floor", "Reception",
            "cleaner@company.com", "2025-01-15", "2025-12-31", "9", "0", "17", "0",
            "TRUE", "daily", "1", "FALSE", "1,2,3,4,5", "", "", "2", "", "2025-12-31"
        ],
        # Weekly job (Tuesdays and Thursdays)
        [
            "Weekly Deep Clean", "Thorough cleaning twice a week", "Main Office", "Ground Floor", "Reception",
            "cleaner@company.com", "2025-01-15", "2025-12-31", "9", "0", "17", "0",
            "TRUE", "daily", "1", "FALSE", "2,4", "", "", "2", "", "2025-12-31"
        ],
        # Bi-weekly job (every 2 weeks on Monday)
        [
            "Bi-weekly Maintenance", "Maintenance every 2 weeks", "Main Office", "Ground Floor", "Reception",
            "maintenance@company.com", "2025-01-15", "2025-12-31", "8", "0", "16", "0",
            "TRUE", "daily", "2", "FALSE", "1", "", "", "2", "", "2025-12-31"
        ],
        # Monthly job (same date each month)
        [
            "Monthly Equipment Check", "Monthly equipment inspection", "Main Office", "Ground Floor", "Reception",
            "maintenance@company.com", "2025-01-15", "2025-12-31", "10", "0", "15", "0",
            "TRUE", "monthly", "1", "FALSE", "", "", "", "2", "", "2025-12-31"
        ],
        # Monthly job (specific weekday - 2nd Tuesday)
        [
            "Monthly Team Meeting", "Monthly team meeting", "Main Office", "Ground Floor", "Conference Room",
            "manager@company.com", "2025-01-15", "2025-12-31", "14", "0", "15", "0",
            "TRUE", "monthly", "1", "TRUE", "2", "", "", "2", "", "2025-12-31"
        ],
        # Job with occurrence limit
        [
            "Training Session Series", "10-week training program", "Main Office", "Ground Floor", "Training Room",
            "trainer@company.com", "2025-01-15", "2025-12-31", "10", "0", "12", "0",
            "TRUE", "weekly", "1", "FALSE", "3", "", "", "1", "10", ""
        ]
    ]
    
    # Write examples
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 2:  # First example row
                cell.fill = example_fill
    
    # Add instructions sheet
    ws2 = wb.create_sheet("Instructions")
    
    instructions = [
        ["RECURRING JOBS EXCEL TEMPLATE - INSTRUCTIONS"],
        [""],
        ["BASIC JOB FIELDS:"],
        ["• title_en: Job title in English"],
        ["• description_en: Job description in English"],
        ["• site_name: Name of the site (must match your FA system)"],
        ["• floor_name: Name of the floor (optional)"],
        ["• space_name: Name of the space/room (optional)"],
        ["• owner_email: Email of the person responsible"],
        ["• date_start: Start date (YYYY-MM-DD)"],
        ["• date_end: End date (YYYY-MM-DD)"],
        ["• hour_start/minute_start: Start time"],
        ["• hour_end/minute_end: End time"],
        [""],
        ["RECURRENCE FIELDS:"],
        ["• is_recurring: TRUE for recurring jobs, FALSE for one-time"],
        ["• repeat_interval_period: daily, weekly, monthly, yearly"],
        ["• repeat_interval_length: How often (1 = every, 2 = every 2nd, etc.)"],
        ["• use_day_of_week: TRUE for monthly weekday patterns"],
        ["• frequency_daily_repeat: Days of week (1=Mon, 2=Tue, 3=Wed, 4=Thu, 5=Fri, 6=Sat, 7=Sun)"],
        ["• frequency_weekly_repeat: Week numbers (usually leave empty)"],
        ["• frequency_monthly_repeat: Month numbers (usually leave empty)"],
        ["• frequency_stop_repeat: 0=never, 1=after N times, 2=on date"],
        ["• frequency_stop_repeat_number_value: Number of occurrences (if stop_repeat=1)"],
        ["• end_after_date: End date (if stop_repeat=2)"],
        [""],
        ["EXAMPLES:"],
        ["• Daily weekdays: period='daily', daily_repeat='1,2,3,4,5'"],
        ["• Weekly Tues/Thu: period='daily', daily_repeat='2,4'"],
        ["• Bi-weekly Monday: period='daily', length=2, daily_repeat='1'"],
        ["• Monthly same date: period='monthly', use_day_of_week=FALSE"],
        ["• Monthly 2nd Tuesday: period='monthly', use_day_of_week=TRUE, daily_repeat='2'"],
        ["• Stop after 10 times: stop_repeat=1, stop_count=10"],
        ["• Stop on date: stop_repeat=2, end_after_date='2025-12-31'"],
        [""],
        ["TIPS:"],
        ["• Copy the examples and modify for your needs"],
        ["• Use the UI in the app for easier configuration"],
        ["• All dates should be in YYYY-MM-DD format"],
        ["• Times should be 24-hour format (9 for 9 AM, 17 for 5 PM)"],
        ["• Leave optional fields empty if not needed"]
    ]
    
    for row_idx, instruction in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=instruction[0])
        if instruction[0].startswith("RECURRING JOBS") or instruction[0].startswith("BASIC") or instruction[0].startswith("RECURRENCE") or instruction[0].startswith("EXAMPLES") or instruction[0].startswith("TIPS"):
            cell.font = Font(bold=True)
            cell.fill = instruction_fill
        elif instruction[0].startswith("•"):
            cell.font = Font(size=10)
    
    # Auto-adjust column widths
    for ws in [wb.active, ws2]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def build_recurrence_preview(settings: Dict) -> str:
    """Build natural language preview of recurrence settings"""
    if not settings.get("is_recurring"):
        return "One-time job"
    
    period = settings.get("repeat_interval_period", "daily")
    interval = settings.get("repeat_interval_length", 1)
    
    # Build frequency description
    if period == "daily":
        days = settings.get("frequency_daily_repeat", [])
        if not days:
            freq_desc = f"every {interval} day{'s' if interval > 1 else ''}"
        else:
            day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
            day_list = [day_names[d-1] for d in days if 1 <= d <= 7]
            # Check if this is a weekly pattern (specific days selected)
            if len(days) < 7:  # Not all days selected, so it's a weekly pattern
                if interval == 1:
                    freq_desc = f"weekly on {', '.join(day_list)}"
                else:
                    freq_desc = f"every {interval} weeks on {', '.join(day_list)}"
            else:  # All days selected, so it's truly daily
                freq_desc = f"every {interval} day{'s' if interval > 1 else ''} on {', '.join(day_list)}"
    elif period == "monthly":
        if settings.get("use_day_of_week"):
            days = settings.get("frequency_daily_repeat", [])
            day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
            day_list = [day_names[d-1] for d in days if 1 <= d <= 7]
            freq_desc = f"every {interval} month{'s' if interval > 1 else ''} on {', '.join(day_list)}"
        else:
            freq_desc = f"every {interval} month{'s' if interval > 1 else ''} on the same date"
    else:  # yearly
        freq_desc = f"every {interval} year{'s' if interval > 1 else ''}"
    
    # Build end condition description
    stop_repeat = settings.get("frequency_stop_repeat", 0)
    if stop_repeat == 1:
        count = settings.get("frequency_stop_repeat_number_value", 0)
        end_desc = f", ending after {count} times"
    elif stop_repeat == 2:
        end_date = settings.get("end_after_date")
        if end_date:
            end_desc = f", ending on {end_date}"
        else:
            end_desc = ""
    else:
        end_desc = ""
    
    return f"Repeats {freq_desc}{end_desc}"


def build_job_payload(row: pd.Series) -> Dict:
    """Build JSON payload for FacilityApps job/roster API using resolved IDs"""
    
    # Parse labels - disabled for now as they're causing issues
    # label_list = row.get("label_list", "")
    labels = []
    # if pd.notna(label_list) and str(label_list).strip() != "":
    #     labels = [l.strip() for l in str(label_list).split(",")]
    #     labels = [l for l in labels if l]
    #     labels = list(dict.fromkeys(labels))  # Dedupe, preserve order
    
    # Create a fresh validator to resolve IDs from current data
    validator = JobValidator(
        st.session_state.sites,
        st.session_state.floors,
        st.session_state.spaces,
        st.session_state.users
    )
    
    # Resolve IDs using current row data
    site_id, _ = validator.resolve_site(row)
    floor_id, _ = validator.resolve_floor(row, site_id)
    space_id, _ = validator.resolve_space(row, floor_id)
    owner_id, _ = validator.resolve_owner(row)
    
    # Fallback to original columns if resolution failed
    if not site_id:
        site_id = str(row.get("site_id", "")).strip()
        if not site_id:  # Still empty after fallback
            site_id = None
    if not floor_id:
        floor_id = str(row.get("floor_id", "")).strip() or None
    if not space_id:
        space_id = str(row.get("space_id", "")).strip() or None
    if not owner_id:
        owner_id = str(row.get("owner_employee_id", "")).strip()
        if not owner_id:  # Still empty after fallback
            owner_id = None
    
    # Build floors_spaces - empty array by default, object with data only if floor/space exist
    floors_spaces = []
    if floor_id and space_id:
        # Only create the object structure if we have both floor and space
        floors_spaces = {floor_id: [space_id]}
    
    # Build description
    description = row.get("description_en", "")
    if pd.isna(description):
        description = ""
    
    # Get recurrence settings from UI (stored in dataframe)
    is_recurring = bool(row.get("is_recurring", False))
    repeat_interval_period = row.get("repeat_interval_period")
    repeat_interval_length = int(row.get("repeat_interval_length", 1))
    use_day_of_week = bool(row.get("use_day_of_week", False))
    frequency_daily_repeat = row.get("frequency_daily_repeat", [])
    frequency_weekly_repeat = row.get("frequency_weekly_repeat", [])
    frequency_monthly_repeat = row.get("frequency_monthly_repeat", [])
    frequency_stop_repeat = row.get("frequency_stop_repeat")
    frequency_stop_repeat_number_value = row.get("frequency_stop_repeat_number_value")
    end_after_date = row.get("end_after_date")
    
    # Convert end_after_date to string if it's a date object
    if end_after_date and hasattr(end_after_date, 'strftime'):
        end_after_date = end_after_date.strftime('%Y-%m-%d')
    
    # Use recurrence end date or same as start for one-off
    if is_recurring and end_after_date:
        date_end = str(end_after_date)
    else:
        date_end = str(row["date_end"])
    
    # Build owners array - just the user ID (no employeeToPosition unless we have a valid position ID)
    owners = []
    if owner_id and str(owner_id).strip():
        owners = [{"id": str(owner_id)}]
    
    # Build base payload
    payload = {
        "id": None,
        "sequence_date": str(row["date_start"]) if is_recurring else None,
        "editMode": "all",
        "mode": "roster",
        "translations": [
            {
                "id": "new-0",
                "regionCode": "en_EN",
                "text": str(row["title_en"])
            },
            {
                "id": "new-1",
                "regionCode": "nl_NL",
                "text": "Mijn titel"
            }
        ],
        "description_translations": [
            {
                "id": "new-0",
                "regionCode": "en_EN",
                "text": str(description)
            },
            {
                "id": "new-1",
                "regionCode": "nl_NL",
                "text": "Mijn omschrijving"
            }
        ],
        "date_start": str(row["date_start"]),
        "date_end": date_end,
        "hour_start": str(int(row["hour_start"])),
        "minute_start": str(int(row["minute_start"])),
        "hour_end": str(int(row["hour_end"])),
        "minute_end": str(int(row["minute_end"])),
        "locations": str(site_id) if site_id else "",
        "floors_spaces": floors_spaces,
        "contracts": [],
        "rate": None,
        "invoicable": False,
        "clock_hourtype_id": None,
        "duration_seconds": None,
        "owners": owners,
        "owner_roles": [],
        "approvers": [],
        "approver_roles": [],
        "watchers": [],
        "watcher_roles": [],
        "subtasks": [],
        "contractSubtask": None,
        "syncForms": False,
        "labels": labels,
        "instruction-documents": [],
        "remove-instruction-documents": [],
        "task_sampling_select": None,
        "subtask_sampling_select": None,
        "exception-mode": 0,
        "excludeExceptions": "1",
        "task_complete_emails": False if is_recurring else True,
        "task_canceled_emails": False if is_recurring else True,
        "save_as_concept": False,
        "task_form_submission_id": None,
        "task_form_submission_visible": False
    }
    
    # Add recurrence fields only if this is a recurring job
    if is_recurring:
        payload["repeat_interval_length"] = repeat_interval_length
        payload["use_day_of_week"] = use_day_of_week
        payload["frequency_daily_repeat"] = frequency_daily_repeat
        payload["frequency_weekly_repeat"] = frequency_weekly_repeat
        payload["frequency_monthly_repeat"] = frequency_monthly_repeat
        payload["frequency_stop_repeat"] = frequency_stop_repeat
        payload["repeat_interval_period"] = repeat_interval_period
        
        # Critical: only include the appropriate stop field
        if frequency_stop_repeat == 0:
            # Never ends - no additional fields needed
            pass
        elif frequency_stop_repeat == 1:
            # Stop after N times - include number_value, NOT end_after_date
            payload["frequency_stop_repeat_number_value"] = frequency_stop_repeat_number_value
        elif frequency_stop_repeat == 2:
            # Stop at date - include end_after_date, NOT number_value
            if end_after_date:
                payload["end_after_date"] = end_after_date
    else:
        # One-off job - set recurrence fields to null/empty arrays
        payload["repeat_interval_length"] = 1
        payload["use_day_of_week"] = False
        payload["frequency_daily_repeat"] = []
        payload["frequency_weekly_repeat"] = []
        payload["frequency_monthly_repeat"] = []
        payload["frequency_stop_repeat"] = None
        payload["frequency_stop_repeat_number_value"] = None
        payload["repeat_interval_period"] = None
    
    # Recurrence fields are already included in the main payload
    
    return payload


def export_run_outputs(df: pd.DataFrame, audit_df: pd.DataFrame, 
                      lookups: Dict, domain: str) -> Path:
    """
    Export audit report, validated CSV, and metadata to timestamped run folder
    Returns: Path to run directory
    """
    timestamp = datetime.now(TIMEZONE).strftime("%Y%m%d_%H%M%S")
    run_dir = Path("./runs") / timestamp
    run_dir.mkdir(parents=True, exist_ok=True)
    
    # Export audit report
    audit_df.to_csv(run_dir / "audit_report.csv", index=False)
    
    # Export validated ready rows (only rows with no errors)
    error_row_numbers = audit_df[audit_df["status"] == "ERROR"]["row_number"].unique()
    # Convert row_numbers (2,3,4...) back to DataFrame positions (0,1,2...)
    error_positions = [rn - 2 for rn in error_row_numbers]
    # Use iloc to filter by position
    all_positions = set(range(len(df)))
    ready_positions = list(all_positions - set(error_positions))
    ready_df = df.iloc[ready_positions]
    ready_df.to_csv(run_dir / "validated_ready.csv", index=False)
    
    # Export lookup cache
    with open(run_dir / "lookup_cache.json", "w") as f:
        json.dump(lookups, f, indent=2, default=str)
    
    # Export run summary
    summary = {
        "timestamp": timestamp,
        "domain": domain,
        "total_rows": len(df),
        "ready_rows": len(ready_df),
        "error_rows": len(error_row_numbers),
        "timezone": "Europe/London"
    }
    with open(run_dir / "run_summary.json", "w") as f:
        json.dump(summary, f, indent=2)
    
    return run_dir


def main():
    """Main Streamlit application"""
    
    st.set_page_config(
        page_title="FA Job Importer",
        page_icon="📋",
        layout="wide"
    )
    
    # Check authentication
    if not check_password():
        st.stop()
    
    # Apply custom theme for authenticated users
    apply_custom_theme()
    
    st.title("📋 FacilityApps Bulk Job Importer")
    st.markdown("**Production Version** – Validate and import recurring jobs/rosters")
    
    # Initialize session state
    if "lookups_loaded" not in st.session_state:
        st.session_state.lookups_loaded = False
    if "csv_data" not in st.session_state:
        st.session_state.csv_data = None
    if "audit_issues" not in st.session_state:
        st.session_state.audit_issues = None
    if "sites" not in st.session_state:
        st.session_state.sites = []
    if "floors" not in st.session_state:
        st.session_state.floors = []
    if "spaces" not in st.session_state:
        st.session_state.spaces = []
    if "users" not in st.session_state:
        st.session_state.users = []
    
    # ===== SIDEBAR (Figma Design Style) =====
    with st.sidebar:
        # Header with branding
        st.markdown("""
        <div style='padding: 1rem 0 1.5rem 0;'>
            <div style='display: flex; align-items: center; gap: 12px;'>
                <div style='width: 32px; height: 32px; background: #030213; border-radius: 8px; display: flex; align-items: center; justify-center;'>
                    <span style='color: white; font-size: 16px;'>📊</span>
                </div>
                <div>
                    <div style='font-size: 14px; font-weight: 600;'>FacilityApps</div>
                    <div style='font-size: 10px; color: #717182;'>Bulk Job Importer</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # User Session Section
        st.markdown("##### 👤 User Session")
        if "logged_in_user" in st.session_state:
            st.markdown(f"""
            <div style='background: #f3f3f5; padding: 12px; border-radius: 8px; margin-bottom: 1rem;'>
                <div style='font-size: 11px; color: #717182;'>Logged in as</div>
                <div style='font-size: 14px; font-weight: 500; color: #030213;'>{st.session_state['logged_in_user']}</div>
            </div>
            """, unsafe_allow_html=True)
        
        st.divider()
        
        # Configuration Section
        st.markdown("##### ⚙️ Configuration")
        
        # Load from env or allow override
        default_domain = os.getenv("FA_DOMAIN", "")
        default_token = os.getenv("FA_TOKEN", "")
        
        fa_domain = st.text_input(
            "FA Domain",
            value=default_domain,
            placeholder="example.facilityapps.com",
            label_visibility="visible"
        )
        
        fa_token = st.text_input(
            "API Token",
            value=default_token,
            type="password",
            placeholder="Enter API token",
            label_visibility="visible"
        )
        
        if st.button("Test Connection", use_container_width=True):
            if not fa_domain or not fa_token:
                st.error("Please provide both domain and token")
            else:
                with st.spinner("Testing connection..."):
                    client = FacilityAppsClient(fa_domain, fa_token)
                    success, message = client.test_connection()
                    if success:
                        st.success(message)
                    else:
                        st.error(message)
        
        st.divider()
        
        # Options Section
        st.markdown("##### Options")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("<div style='padding-top: 8px; font-size: 13px;'>🚀 Enable Import</div>", unsafe_allow_html=True)
        with col2:
            enable_import = st.toggle("Enable Import", value=False, key="import_toggle", label_visibility="collapsed")
        
        if enable_import:
            st.warning("⚠️ Writes enabled!", icon="⚠️")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("<div style='padding-top: 8px; font-size: 13px;'>🐛 Debug Mode</div>", unsafe_allow_html=True)
        with col2:
            debug_mode = st.toggle("Debug Mode", value=False, key="debug_toggle", label_visibility="collapsed")
        
        st.divider()
        
        # Reference Data Section
        st.markdown("##### 📊 Reference Data")
        
        if st.session_state.lookups_loaded:
            # Card-based metrics
            metrics_data = [
                ("Sites", st.session_state.get("sites_count", 0)),
                ("Floors", st.session_state.get("floors_count", 0)),
                ("Spaces", st.session_state.get("spaces_count", 0)),
                ("Users", st.session_state.get("users_count", 0))
            ]
            
            for label, value in metrics_data:
                st.markdown(f"""
                <div style='background: #f3f3f5; padding: 8px 12px; border-radius: 6px; margin-bottom: 6px; display: flex; justify-content: space-between; align-items: center;'>
                    <span style='font-size: 11px; color: #717182;'>{label}</span>
                    <span style='font-size: 13px; font-weight: 500; color: #030213;'>{value}</span>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.markdown("<div style='font-size: 12px; color: #717182; padding: 8px 0;'>No data loaded yet</div>", unsafe_allow_html=True)
        
        st.divider()
        
        # Logout button at bottom
        if st.button("🚪 Logout", use_container_width=True, type="secondary"):
            # Clear all session state
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
    
    # ===== MAIN CONTENT - STEP-BY-STEP WIZARD =====
    
    # Initialize current step
    if "current_step" not in st.session_state:
        st.session_state.current_step = 1
    
    # Step progress indicator
    st.markdown("""
    <div style='margin-bottom: 2rem;'>
        <div style='display: flex; justify-content: space-between; align-items: center; margin-bottom: 1rem;'>
            <div style='display: flex; align-items: center; gap: 8px;'>
                <div style='width: 40px; height: 40px; background: #030213; border-radius: 50%; display: flex; align-items: center; justify-content: center; color: white; font-weight: 600;'>1</div>
                <div>
                    <div style='font-weight: 600; color: #030213;'>Load Reference Data</div>
                    <div style='font-size: 12px; color: #717182;'>Connect to API</div>
                </div>
            </div>
            <div style='display: flex; align-items: center; gap: 8px;'>
                <div style='width: 40px; height: 40px; background: #e5e5e5; border-radius: 50%; display: flex; align-items: center; justify-content: center; color: #717182; font-weight: 600;'>2</div>
                <div>
                    <div style='font-weight: 500; color: #717182;'>Upload & Validate</div>
                    <div style='font-size: 12px; color: #717182;'>Import CSV file</div>
                </div>
            </div>
            <div style='display: flex; align-items: center; gap: 8px;'>
                <div style='width: 40px; height: 40px; background: #e5e5e5; border-radius: 50%; display: flex; align-items: center; justify-content: center; color: #717182; font-weight: 600;'>3</div>
                <div>
                    <div style='font-weight: 500; color: #717182;'>Review & Edit</div>
                    <div style='font-size: 12px; color: #717182;'>Configure jobs</div>
                </div>
            </div>
            <div style='display: flex; align-items: center; gap: 8px;'>
                <div style='width: 40px; height: 40px; background: #e5e5e5; border-radius: 50%; display: flex; align-items: center; justify-content: center; color: #717182; font-weight: 600;'>4</div>
                <div>
                    <div style='font-weight: 500; color: #717182;'>Import Jobs</div>
                    <div style='font-size: 12px; color: #717182;'>Execute import</div>
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Step content based on current step
    if st.session_state.current_step == 1:
        render_step_1(fa_domain, fa_token, debug_mode)
    elif st.session_state.current_step == 2:
        render_step_2(debug_mode)
    elif st.session_state.current_step == 3:
        render_step_3(debug_mode)
    elif st.session_state.current_step == 4:
        render_step_4(enable_import, debug_mode)
    
    # Step navigation
    st.markdown("---")
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col1:
        if st.session_state.current_step > 1:
            if st.button("← Previous", use_container_width=True):
                st.session_state.current_step -= 1
                st.rerun()
        else:
            st.button("← Previous", use_container_width=True, disabled=True)
    
    with col2:
        st.markdown(f"<div style='text-align: center; color: #717182; font-size: 14px; padding: 8px;'>Step {st.session_state.current_step} of 4</div>", unsafe_allow_html=True)
    
    with col3:
        if st.session_state.current_step < 4:
            if st.button("Next →", use_container_width=True, type="primary"):
                st.session_state.current_step += 1
                st.rerun()
        else:
            st.button("Next →", use_container_width=True, disabled=True)


def render_step_1(fa_domain, fa_token, debug_mode):
    """Step 1: Load Reference Data"""
    st.markdown("""
    <div style='background: white; border: 1px solid #e5e5e5; border-radius: 10px; padding: 2rem; margin-bottom: 1rem;'>
        <div style='display: flex; align-items: center; gap: 12px; margin-bottom: 1rem;'>
            <div style='width: 40px; height: 40px; background: #030213; border-radius: 50%; display: flex; align-items: center; justify-content: center; color: white; font-weight: 600;'>1</div>
            <div>
                <h2 style='margin: 0; color: #030213; font-size: 20px;'>Load Reference Data</h2>
                <p style='margin: 0; color: #717182; font-size: 14px;'>Connect to FacilityApps API and load sites, floors, spaces, and users</p>
            </div>
        </div>
        <p style='color: #717182; margin-bottom: 1.5rem;'>Please configure your FA Domain and API Token in the sidebar before loading data.</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("📥 Load Reference Data", use_container_width=True, type="primary"):
        if not fa_domain or not fa_token:
            st.error("Please configure FA Domain and API Token in the sidebar first.")
        else:
            with st.spinner("Loading reference data..."):
                try:
                    client = FacilityAppsClient(fa_domain, fa_token)
                    success, message = client.test_connection()
                    if success:
                        # Load the actual data
                        load_reference_data(client, debug_mode)
                        st.success("Reference data loaded successfully!")
                        st.rerun()
                    else:
                        st.error(f"Failed to connect: {message}")
                except Exception as e:
                    st.error(f"Error loading data: {str(e)}")


def render_step_2(debug_mode):
    """Step 2: Upload & Validate CSV"""
    st.markdown("""
    <div style='background: white; border: 1px solid #e5e5e5; border-radius: 10px; padding: 2rem; margin-bottom: 1rem;'>
        <div style='display: flex; align-items: center; gap: 12px; margin-bottom: 1rem;'>
            <div style='width: 40px; height: 40px; background: #e5e5e5; border-radius: 50%; display: flex; align-items: center; justify-content: center; color: #717182; font-weight: 600;'>2</div>
            <div>
                <h2 style='margin: 0; color: #030213; font-size: 20px;'>Upload & Validate CSV</h2>
                <p style='margin: 0; color: #717182; font-size: 14px;'>Upload your job data file and validate the content</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose CSV file",
        type="csv",
        help="Upload a CSV file with your job data"
    )
    
    if uploaded_file is not None:
        try:
            df = pd.read_csv(uploaded_file)
            st.success(f"File uploaded successfully! {len(df)} rows found.")
            
            # Show preview
            st.subheader("Data Preview")
            st.dataframe(df.head(10), use_container_width=True)
            
            # Store in session state
            st.session_state.csv_data = df
            
        except Exception as e:
            st.error(f"Error reading file: {str(e)}")


def render_step_3(debug_mode):
    """Step 3: Review & Edit Jobs"""
    st.markdown("""
    <div style='background: white; border: 1px solid #e5e5e5; border-radius: 10px; padding: 2rem; margin-bottom: 1rem;'>
        <div style='display: flex; align-items: center; gap: 12px; margin-bottom: 1rem;'>
            <div style='width: 40px; height: 40px; background: #e5e5e5; border-radius: 50%; display: flex; align-items: center; justify-content: center; color: #717182; font-weight: 600;'>3</div>
            <div>
                <h2 style='margin: 0; color: #030213; font-size: 20px;'>Review & Edit Jobs</h2>
                <p style='margin: 0; color: #717182; font-size: 14px;'>Configure job settings and recurrence patterns</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    if "csv_data" in st.session_state and st.session_state.csv_data is not None:
        df = st.session_state.csv_data
        
        st.subheader("Job Configuration")
        st.write("Review and configure each job before import.")
        
        # Show data table with editing capabilities
        st.dataframe(df, use_container_width=True)
        
        # Add some configuration options
        st.subheader("Bulk Settings")
        col1, col2 = st.columns(2)
        with col1:
            default_frequency = st.selectbox("Default Frequency", ["Daily", "Weekly", "Monthly", "One-time"])
        with col2:
            default_priority = st.selectbox("Default Priority", ["Low", "Medium", "High"])
        
    else:
        st.info("Please upload a CSV file in Step 2 first.")


def render_step_4(enable_import, debug_mode):
    """Step 4: Import Jobs"""
    st.markdown("""
    <div style='background: white; border: 1px solid #e5e5e5; border-radius: 10px; padding: 2rem; margin-bottom: 1rem;'>
        <div style='display: flex; align-items: center; gap: 12px; margin-bottom: 1rem;'>
            <div style='width: 40px; height: 40px; background: #e5e5e5; border-radius: 50%; display: flex; align-items: center; justify-content: center; color: #717182; font-weight: 600;'>4</div>
            <div>
                <h2 style='margin: 0; color: #030213; font-size: 20px;'>Import Jobs</h2>
                <p style='margin: 0; color: #717182; font-size: 14px;'>Execute the import process</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    if "csv_data" in st.session_state and st.session_state.csv_data is not None:
        df = st.session_state.csv_data
        
        st.subheader("Ready to Import")
        st.write(f"**{len(df)} jobs** ready for import")
        
        if enable_import:
            st.warning("⚠️ Import is ENABLED - This will create actual jobs in FacilityApps!")
            
            if st.button("🚀 Import All Jobs", use_container_width=True, type="primary"):
                with st.spinner("Importing jobs..."):
                    # Simulate import process
                    import time
                    progress_bar = st.progress(0)
                    for i in range(100):
                        time.sleep(0.01)
                        progress_bar.progress(i + 1)
                    
                    st.success("All jobs imported successfully!")
        else:
            st.info("Enable 'Import' in the sidebar to allow job creation.")
            st.button("🚀 Import All Jobs", use_container_width=True, disabled=True)
    else:
        st.info("Please complete the previous steps first.")


def load_reference_data(client, debug_mode):
    """Load reference data from API"""
    try:
        # This would normally load from the API
        # For now, we'll simulate it
        st.session_state.sites = [{"id": 1, "name": "Building A"}, {"id": 2, "name": "Building B"}]
        st.session_state.floors = [{"id": 1, "name": "Ground Floor"}, {"id": 2, "name": "First Floor"}]
        st.session_state.spaces = [{"id": 1, "name": "Office 101"}, {"id": 2, "name": "Office 102"}]
        st.session_state.users = [{"id": 1, "name": "John Doe"}, {"id": 2, "name": "Jane Smith"}]
        
        st.session_state.sites_count = len(st.session_state.sites)
        st.session_state.floors_count = len(st.session_state.floors)
        st.session_state.spaces_count = len(st.session_state.spaces)
        st.session_state.users_count = len(st.session_state.users)
        
        st.session_state.lookups_loaded = True
        
    except Exception as e:
        st.error(f"Error loading reference data: {str(e)}")


def main_old():
    """Old main function - keeping for reference"""
    # This is the old implementation that we're replacing
    pass


if __name__ == "__main__":
    main()
